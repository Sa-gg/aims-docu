"""
Build AIMS Gantt Chart
Copies the ATLAS Gantt Chart exactly, changes only text labels (objectives, title, phase tasks),
and keeps ALL dates and formatting identical.
"""

import openpyxl
import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SRC = "ganttchart/ATLAS_GanttChart_revised.xlsx"
DST = "ganttchart/draft/AIMS_GanttChart_v1.xlsx"

# ─────────────────────────────────────────────
# STEP 0 — Record ALL original merges BEFORE insert
# ─────────────────────────────────────────────
wb_ref = load_workbook(SRC)
ws_ref = wb_ref.active
# Capture every merge that starts at row >= 7 (will be affected by insert at row 7)
orig_affected_merges = [
    (m.min_row, m.min_col, m.max_row, m.max_col)
    for m in ws_ref.merged_cells.ranges
    if m.min_row >= 7
]
wb_ref.close()

# ─────────────────────────────────────────────
# Load working copy
# ─────────────────────────────────────────────
wb = load_workbook(SRC)
ws = wb.active

# ─────────────────────────────────────────────
# STEP 1 — Objectives section (rows 1-6, no shift needed)
# ─────────────────────────────────────────────
ws["B2"] = "1. design and develop a web-based educational platform with the following technical features:"
ws["C3"] = "1.1. Centralized learning portal with material distribution and task submission."
ws["C4"] = "1.2. AI-Powered Quiz Generator with instructor validation and approval controls."
ws["C5"] = "1.3. Automated assessment and grading engine supporting teacher validation."
ws["C6"] = "1.4. Dynamic remedial quiz generator for targeted student intervention."

# ─────────────────────────────────────────────
# STEP 2 — Insert new row 7 for Objective 1.5
# ─────────────────────────────────────────────
# Save C6 formatting to apply to C7
c6 = ws["C6"]
c6_font = copy.copy(c6.font)
c6_fill = copy.copy(c6.fill)
c6_alignment = copy.copy(c6.alignment)
c6_border = copy.copy(c6.border)
c6_num_format = c6.number_format

ws.insert_rows(7)

# ─────────────────────────────────────────────
# STEP 3 — Fix ALL stale merges caused by insert_rows
# openpyxl does not always remove old merge references after insert_rows.
# Strategy: remove ALL merges that touch rows >= 7 (both stale and new),
# then re-add ALL original affected merges shifted by +1.
# ─────────────────────────────────────────────
# Collect current merges at row >= 7 (includes stale duplicates)
current_affected = [m.coord for m in ws.merged_cells.ranges if m.min_row >= 7]
for coord in current_affected:
    try:
        ws.unmerge_cells(coord)
    except Exception:
        pass

# Re-add every original affected merge, shifted +1
for min_row, min_col, max_row, max_col in orig_affected_merges:
    start_col = get_column_letter(min_col)
    end_col = get_column_letter(max_col)
    ws.merge_cells(f"{start_col}{min_row + 1}:{end_col}{max_row + 1}")

# ─────────────────────────────────────────────
# STEP 4 — Populate the new row 7 (Objective 1.5)
# ─────────────────────────────────────────────
ws["C7"] = "1.5. Mastery-based learning module with configurable progression locks."
ws.merge_cells("C7:O7")
c7 = ws["C7"]
c7.font = c6_font
c7.fill = c6_fill
c7.alignment = c6_alignment
c7.border = c6_border
c7.number_format = c6_num_format

# ─────────────────────────────────────────────
# STEP 5 — Main title row (original row 12 → now row 13)
# ─────────────────────────────────────────────
ws["A13"] = (
    "A. I. M. S. (Automated Intervention and Mastery System): "
    "A Web-Based Educational Platform with AI-Driven Assessment and Dynamic Remediation"
)

# ─────────────────────────────────────────────
# STEP 6 — Gantt task text replacements
# All original row numbers ≥ 7 are now +1
# ─────────────────────────────────────────────

# Row 21 — was "Initial Client Meetings (EMEMHS & HNHS) & Adviser Acceptance"
ws["B21"] = "Initial Beneficiary Meetings & Adviser Acceptance"

# Row 24 — was "Demonstrate: Requirements Interview & Skeleton Prototype Presentation to beneficiary (HNHS)"
ws["B24"] = "Demonstrate: Requirements Interview & Skeleton Prototype Presentation to Beneficiary"

# Row 28 — Build Iteration 1: ATLAS-specific algorithm label
ws["C28"] = "Build: Initial LMS Portal & AI Quiz Module Prototyping"

# Row 44 — "Live Website" → "Live Web Application"
ws["C44"] = "Demonstrate: 100% System Presentation & Live Web Application Launch"

# ─────────────────────────────────────────────
# STEP 7 — Save
# ─────────────────────────────────────────────
wb.save(DST)
print(f"Saved: {DST}")
